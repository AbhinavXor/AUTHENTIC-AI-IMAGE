import asyncio
import csv
import hashlib
import logging
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import (
    date,
    datetime,
    time,
)
from io import (
    BytesIO,
    StringIO,
)
from pathlib import PurePosixPath
from time import monotonic
from typing import Any
from zipfile import (
    BadZipFile,
    ZipFile,
    is_zipfile,
)

from google import genai
from google.genai import (
    errors,
    types,
)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.gemini_settings import (
    gemini_settings,
)
from core.spreadsheet_settings import (
    spreadsheet_settings,
)
from schemas.chat import TokenUsage
from schemas.spreadsheets import (
    SpreadsheetCitation,
    SpreadsheetSourceKind,
    SpreadsheetType,
)


logger = logging.getLogger(__name__)


SPREADSHEET_SYSTEM_PROMPT = """
You are Serenya Spreadsheet Intelligence.

You receive deterministic profiles and selected row ranges
extracted from one CSV or XLSX spreadsheet.

NON-NEGOTIABLE RULES

- Answer only from the supplied spreadsheet sources.
- Do not invent sheets, rows, columns, cells, totals, formulas,
  categories, trends, anomalies, or missing values.
- Treat the supplied profile statistics as deterministic results.
- Do not recalculate a supplied statistic differently.
- Do not claim a complete-workbook conclusion when the response
  says the workbook was truncated.
- Formula text may be inspected, but formulas were not executed.
- Cached formula values may be absent or stale.
- Cite factual claims using the exact supplied source labels.
- Put every source in separate brackets.
- Example:
  [Sheet Revenue profile] [Sheet Revenue rows 2-20]
- Never cite a source that was not supplied.
- If the selected sources do not support the answer, say:
  "The supplied spreadsheet sources do not provide enough
  information to answer this."
- Respond in the user's language whenever practical.
- Return clean Markdown.
- Begin directly with the answer.
- Never reveal hidden reasoning or private chain-of-thought.
""".strip()


_STOP_WORDS = frozenset(
    {
        "about",
        "all",
        "also",
        "analyze",
        "and",
        "answer",
        "are",
        "average",
        "csv",
        "data",
        "excel",
        "file",
        "find",
        "from",
        "give",
        "have",
        "into",
        "main",
        "more",
        "most",
        "please",
        "provide",
        "sheet",
        "show",
        "spreadsheet",
        "summarize",
        "summary",
        "table",
        "that",
        "the",
        "their",
        "this",
        "what",
        "when",
        "where",
        "which",
        "with",
        "workbook",
        "xlsx",
    }
)


class SpreadsheetValidationError(
    RuntimeError
):
    """Raised for unsafe or invalid spreadsheets."""


class SpreadsheetConfigurationError(
    RuntimeError
):
    """Raised when the analysis model is unavailable."""


class SpreadsheetResponseError(
    RuntimeError
):
    """Raised when no usable answer is produced."""


@dataclass(frozen=True, slots=True)
class SpreadsheetSource:
    source_id: str
    label: str
    kind: SpreadsheetSourceKind
    text: str
    character_count: int


@dataclass(frozen=True, slots=True)
class ExtractedSpreadsheet:
    spreadsheet_type: SpreadsheetType

    sheet_names: tuple[str, ...]
    sources: tuple[
        SpreadsheetSource,
        ...
    ]

    rows_scanned: int
    maximum_columns_seen: int
    formula_count: int

    truncated: bool


@dataclass(frozen=True, slots=True)
class SpreadsheetAnalysis:
    answer: str
    model: str
    request_id: str | None
    usage: TokenUsage

    spreadsheet_type: SpreadsheetType

    sheet_names: tuple[str, ...]
    rows_scanned: int
    maximum_columns_seen: int
    formula_count: int

    truncated: bool

    selected_sources: tuple[str, ...]
    citations: tuple[
        SpreadsheetCitation,
        ...
    ]


def _safe_label(
    value: str,
) -> str:
    normalized = re.sub(
        r"[\[\]\r\n]+",
        " ",
        value,
    )

    normalized = re.sub(
        r"\s+",
        " ",
        normalized,
    ).strip()

    return normalized[:100] or "Sheet"


def _cell_text(
    value: Any,
) -> str:
    if value is None:
        return ""

    if isinstance(
        value,
        datetime,
    ):
        return value.isoformat(
            sep=" "
        )

    if isinstance(
        value,
        (
            date,
            time,
        ),
    ):
        return value.isoformat()

    if isinstance(
        value,
        float,
    ):
        if not math.isfinite(value):
            return str(value)

        return format(
            value,
            ".12g",
        )

    normalized = str(value)

    normalized = normalized.replace(
        "\r\n",
        "\n",
    ).replace(
        "\r",
        "\n",
    )

    normalized = re.sub(
        r"\s+",
        " ",
        normalized,
    ).strip()

    return normalized[
        :spreadsheet_settings
        .maximum_cell_characters
    ]


def _is_numeric(
    value: Any,
) -> bool:
    return (
        isinstance(
            value,
            (
                int,
                float,
            ),
        )
        and not isinstance(
            value,
            bool,
        )
        and (
            not isinstance(
                value,
                float,
            )
            or math.isfinite(value)
        )
    )


def _number_text(
    value: float,
) -> str:
    return format(
        value,
        ".12g",
    )


def _decode_csv(
    file_bytes: bytes,
) -> str:
    if b"\x00" in file_bytes[:4_096]:
        if file_bytes.startswith(
            (
                b"\xff\xfe",
                b"\xfe\xff",
            )
        ):
            try:
                return file_bytes.decode(
                    "utf-16"
                )
            except UnicodeDecodeError as error:
                raise SpreadsheetValidationError(
                    "The CSV uses an unsupported encoding."
                ) from error

        raise SpreadsheetValidationError(
            "The uploaded CSV appears to contain binary data."
        )

    try:
        return file_bytes.decode(
            "utf-8-sig"
        )

    except UnicodeDecodeError as error:
        raise SpreadsheetValidationError(
            "CSV files must use UTF-8 or UTF-16 encoding."
        ) from error


def _validate_xlsx_archive(
    file_bytes: bytes,
) -> None:
    stream = BytesIO(
        file_bytes
    )

    if not is_zipfile(stream):
        raise SpreadsheetValidationError(
            "The uploaded file is not a valid XLSX archive."
        )

    stream.seek(0)

    try:
        with ZipFile(stream) as archive:
            entries = archive.infolist()

            if (
                len(entries)
                > spreadsheet_settings
                .maximum_xlsx_archive_entries
            ):
                raise SpreadsheetValidationError(
                    "The XLSX archive contains too many entries."
                )

            total_uncompressed = 0
            entry_names: set[str] = set()

            for entry in entries:
                entry_path = PurePosixPath(
                    entry.filename
                )

                if (
                    entry_path.is_absolute()
                    or ".." in entry_path.parts
                ):
                    raise SpreadsheetValidationError(
                        "The XLSX archive contains an unsafe path."
                    )

                if entry.flag_bits & 0x1:
                    raise SpreadsheetValidationError(
                        "Encrypted XLSX archives are not supported."
                    )

                total_uncompressed += (
                    entry.file_size
                )

                if (
                    total_uncompressed
                    > spreadsheet_settings
                    .maximum_xlsx_uncompressed_bytes
                ):
                    raise SpreadsheetValidationError(
                        "The expanded XLSX exceeds the safety limit."
                    )

                normalized_name = (
                    entry.filename.lower()
                )

                if (
                    normalized_name.endswith(
                        "vbaproject.bin"
                    )
                    or "/macrosheets/"
                    in normalized_name
                ):
                    raise SpreadsheetValidationError(
                        "Macro-enabled workbooks are not supported."
                    )

                entry_names.add(
                    entry.filename
                )

            required_entries = {
                "[Content_Types].xml",
                "xl/workbook.xml",
            }

            if not required_entries.issubset(
                entry_names
            ):
                raise SpreadsheetValidationError(
                    "The archive does not contain a valid XLSX workbook."
                )

    except BadZipFile as error:
        raise SpreadsheetValidationError(
            "The XLSX archive is corrupted."
        ) from error


def _query_terms(
    prompt: str,
) -> frozenset[str]:
    return frozenset(
        term
        for term in re.findall(
            r"[a-zA-Z0-9][a-zA-Z0-9_-]{2,}",
            prompt.lower(),
        )
        if term not in _STOP_WORDS
    )


def _looks_like_broad_request(
    prompt: str,
) -> bool:
    lowered = prompt.lower()

    broad_terms = (
        "summarize",
        "summary",
        "overview",
        "entire workbook",
        "whole workbook",
        "all sheets",
        "main trends",
        "key findings",
        "data quality",
        "missing values",
        "duplicates",
    )

    return any(
        term in lowered
        for term in broad_terms
    )


def _source_score(
    source: SpreadsheetSource,
    terms: frozenset[str],
) -> float:
    if not terms:
        return 0.0

    searchable = (
        f"{source.label}\n"
        f"{source.text}"
    ).lower()

    score = 0.0

    for term in terms:
        occurrences = searchable.count(
            term
        )

        if occurrences:
            score += min(
                occurrences,
                20,
            ) * 2.0

        if term in source.label.lower():
            score += 6.0

    if source.kind == "profile":
        score += 1.0

    return score


def _evenly_sample(
    sources: tuple[
        SpreadsheetSource,
        ...
    ],
    count: int,
) -> list[
    SpreadsheetSource
]:
    if count <= 0 or not sources:
        return []

    if count >= len(sources):
        return list(sources)

    if count == 1:
        return [
            sources[0],
        ]

    final_index = len(
        sources
    ) - 1

    indexes = {
        round(
            position
            * final_index
            / (count - 1)
        )
        for position in range(
            count
        )
    }

    return [
        sources[index]
        for index in sorted(
            indexes
        )
    ]


def _row_signature(
    values: list[str],
) -> bytes:
    serialized = "\x1f".join(
        values
    ).encode(
        "utf-8",
        errors="replace",
    )

    return hashlib.sha256(
        serialized
    ).digest()


def _profile_text(
    *,
    name: str,
    headers: list[str],
    data_rows: int,
    columns_seen: int,
    formula_count: int,
    duplicate_rows: int,
    missing_counts: dict[
        int,
        int,
    ],
    numeric_stats: dict[
        int,
        dict[str, float],
    ],
    truncated: bool,
) -> str:
    lines = [
        f"Dataset: {name}",
        f"Data rows scanned: {data_rows}",
        f"Columns seen: {columns_seen}",
        f"Formula cells found: {formula_count}",
        f"Duplicate data rows found: {duplicate_rows}",
        f"Truncated during scanning: {'yes' if truncated else 'no'}",
    ]

    missing_parts: list[str] = []

    for index in range(
        min(
            len(headers),
            spreadsheet_settings
            .maximum_context_columns,
        )
    ):
        missing = missing_counts.get(
            index,
            0,
        )

        if missing <= 0:
            continue

        missing_parts.append(
            f"{headers[index]}={missing}"
        )

    if missing_parts:
        lines.append(
            "Missing values by column: "
            + "; ".join(
                missing_parts
            )
        )

    numeric_lines: list[str] = []

    for index in sorted(
        numeric_stats
    ):
        if index >= len(headers):
            continue

        stats = numeric_stats[
            index
        ]

        count = int(
            stats["count"]
        )

        if count <= 0:
            continue

        average = (
            stats["sum"]
            / count
        )

        numeric_lines.append(
            (
                f"{headers[index]}: "
                f"count={count}, "
                f"sum={_number_text(stats['sum'])}, "
                f"average={_number_text(average)}, "
                f"minimum={_number_text(stats['min'])}, "
                f"maximum={_number_text(stats['max'])}"
            )
        )

    if numeric_lines:
        lines.append(
            "Numeric column statistics:"
        )

        lines.extend(
            f"- {line}"
            for line in numeric_lines[
                :spreadsheet_settings
                .maximum_context_columns
            ]
        )

    return "\n".join(
        lines
    )


class SpreadsheetService:
    provider_name = "gemini"

    def __init__(self) -> None:
        self._client: (
            genai.Client
            | None
        ) = None

    def _get_client(
        self,
    ) -> genai.Client:
        if not gemini_settings.api_key:
            raise SpreadsheetConfigurationError(
                "Gemini spreadsheet analysis is not configured."
            )

        if self._client is None:
            self._client = genai.Client(
                api_key=gemini_settings.api_key,
                http_options=types.HttpOptions(
                    timeout=int(
                        gemini_settings.timeout_seconds
                        * 1000
                    ),
                ),
            )

        return self._client

    @staticmethod
    def _model_candidates(
    ) -> tuple[str, ...]:
        return tuple(
            dict.fromkeys(
                model
                for model in (
                    gemini_settings.quality_model,
                    gemini_settings.fallback_model,
                    gemini_settings.preview_model,
                    gemini_settings.fast_model,
                )
                if model
            )
        )

    @staticmethod
    def _extract_csv_sync(
        file_bytes: bytes,
    ) -> ExtractedSpreadsheet:
        started_at = monotonic()

        decoded = _decode_csv(
            file_bytes
        )

        sample = decoded[:16_384]

        try:
            dialect = csv.Sniffer().sniff(
                sample,
                delimiters=",;\t|",
            )

        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(
            StringIO(
                decoded,
                newline="",
            ),
            dialect,
        )

        sources: list[
            SpreadsheetSource
        ] = []

        headers: list[str] = []

        rows_scanned = 0
        data_rows = 0
        maximum_columns_seen = 0
        duplicate_rows = 0
        truncated = False

        missing_counts: dict[
            int,
            int,
        ] = defaultdict(int)

        numeric_stats: dict[
            int,
            dict[str, float],
        ] = {}

        row_hashes: set[bytes] = set()

        chunk_lines: list[str] = []
        chunk_start = 0
        chunk_end = 0

        def flush_chunk() -> None:
            nonlocal chunk_lines
            nonlocal chunk_start
            nonlocal chunk_end

            if not chunk_lines:
                return

            label = (
                f"CSV rows "
                f"{chunk_start}-{chunk_end}"
            )

            text = "\n".join(
                chunk_lines
            )

            sources.append(
                SpreadsheetSource(
                    source_id=(
                        f"csv-rows-"
                        f"{chunk_start}-"
                        f"{chunk_end}"
                    ),
                    label=label,
                    kind="csv_rows",
                    text=text,
                    character_count=len(
                        text
                    ),
                )
            )

            chunk_lines = []
            chunk_start = 0
            chunk_end = 0

        try:
            for row_index, row in enumerate(
                reader,
                start=1,
            ):
                if (
                    monotonic()
                    - started_at
                    > spreadsheet_settings
                    .maximum_parse_seconds
                ):
                    raise SpreadsheetValidationError(
                        "The CSV took too long to parse safely."
                    )

                if (
                    rows_scanned
                    >= spreadsheet_settings
                    .maximum_total_rows
                ):
                    truncated = True
                    break

                limited_row = row[
                    :spreadsheet_settings
                    .maximum_columns
                ]

                values = [
                    _cell_text(
                        value
                    )
                    for value in limited_row
                ]

                if not any(values):
                    continue

                rows_scanned += 1

                maximum_columns_seen = max(
                    maximum_columns_seen,
                    len(values),
                )

                if not headers:
                    headers = [
                        value
                        or (
                            "Column "
                            f"{get_column_letter(index)}"
                        )
                        for index, value in enumerate(
                            values,
                            start=1,
                        )
                    ]

                    continue

                data_rows += 1

                while len(headers) < len(values):
                    headers.append(
                        "Column "
                        f"{get_column_letter(len(headers) + 1)}"
                    )

                padded_values = (
                    values
                    + [""] * (
                        len(headers)
                        - len(values)
                    )
                )

                signature = _row_signature(
                    padded_values[
                        :spreadsheet_settings
                        .maximum_context_columns
                    ]
                )

                if signature in row_hashes:
                    duplicate_rows += 1
                else:
                    row_hashes.add(
                        signature
                    )

                display_parts: list[str] = []

                for column_index, value in enumerate(
                    padded_values
                ):
                    if column_index >= (
                        spreadsheet_settings
                        .maximum_context_columns
                    ):
                        break

                    header = headers[
                        column_index
                    ]

                    if not value:
                        missing_counts[
                            column_index
                        ] += 1

                    else:
                        try:
                            number = float(
                                value.replace(
                                    ",",
                                    "",
                                )
                            )

                        except ValueError:
                            number = None

                        if (
                            number is not None
                            and math.isfinite(
                                number
                            )
                        ):
                            stats = (
                                numeric_stats
                                .setdefault(
                                    column_index,
                                    {
                                        "count": 0.0,
                                        "sum": 0.0,
                                        "min": number,
                                        "max": number,
                                    },
                                )
                            )

                            stats["count"] += 1
                            stats["sum"] += number
                            stats["min"] = min(
                                stats["min"],
                                number,
                            )
                            stats["max"] = max(
                                stats["max"],
                                number,
                            )

                    display_parts.append(
                        f"{header}={value or '∅'}"
                    )

                if not chunk_lines:
                    chunk_start = row_index

                chunk_end = row_index

                chunk_lines.append(
                    (
                        f"Row {row_index}: "
                        + "; ".join(
                            display_parts
                        )
                    )
                )

                if (
                    len(chunk_lines)
                    >= spreadsheet_settings
                    .row_chunk_size
                ):
                    flush_chunk()

                    if (
                        len(sources)
                        >= spreadsheet_settings
                        .maximum_sources
                    ):
                        truncated = True
                        break

        except csv.Error as error:
            raise SpreadsheetValidationError(
                "The CSV structure is malformed."
            ) from error

        flush_chunk()

        if not headers:
            raise SpreadsheetValidationError(
                "The CSV does not contain readable rows."
            )

        profile = _profile_text(
            name="CSV",
            headers=headers,
            data_rows=data_rows,
            columns_seen=maximum_columns_seen,
            formula_count=0,
            duplicate_rows=duplicate_rows,
            missing_counts=dict(
                missing_counts
            ),
            numeric_stats=numeric_stats,
            truncated=truncated,
        )

        sources.insert(
            0,
            SpreadsheetSource(
                source_id="csv-profile",
                label="CSV profile",
                kind="profile",
                text=profile,
                character_count=len(
                    profile
                ),
            ),
        )

        return ExtractedSpreadsheet(
            spreadsheet_type="csv",
            sheet_names=("CSV",),
            sources=tuple(
                sources
            ),
            rows_scanned=rows_scanned,
            maximum_columns_seen=(
                maximum_columns_seen
            ),
            formula_count=0,
            truncated=truncated,
        )

    @staticmethod
    def _extract_xlsx_sync(
        file_bytes: bytes,
    ) -> ExtractedSpreadsheet:
        started_at = monotonic()

        _validate_xlsx_archive(
            file_bytes
        )

        formula_stream = BytesIO(
            file_bytes
        )

        value_stream = BytesIO(
            file_bytes
        )

        try:
            formula_workbook = load_workbook(
                formula_stream,
                read_only=True,
                data_only=False,
                keep_links=False,
            )

            value_workbook = load_workbook(
                value_stream,
                read_only=True,
                data_only=True,
                keep_links=False,
            )

        except Exception as error:
            raise SpreadsheetValidationError(
                "The XLSX workbook could not be opened."
            ) from error

        try:
            sheet_names = tuple(
                formula_workbook.sheetnames
            )

            if not sheet_names:
                raise SpreadsheetValidationError(
                    "The workbook does not contain any worksheets."
                )

            if (
                len(sheet_names)
                > spreadsheet_settings
                .maximum_worksheets
            ):
                raise SpreadsheetValidationError(
                    "The workbook contains more than "
                    f"{spreadsheet_settings.maximum_worksheets} worksheets."
                )

            sources: list[
                SpreadsheetSource
            ] = []

            total_rows_scanned = 0
            maximum_columns_seen = 0
            total_formula_count = 0
            workbook_truncated = False

            for sheet_name in sheet_names:
                if (
                    monotonic()
                    - started_at
                    > spreadsheet_settings
                    .maximum_parse_seconds
                ):
                    raise SpreadsheetValidationError(
                        "The workbook took too long to parse safely."
                    )

                formula_sheet = (
                    formula_workbook[
                        sheet_name
                    ]
                )

                value_sheet = (
                    value_workbook[
                        sheet_name
                    ]
                )

                safe_sheet_name = (
                    _safe_label(
                        sheet_name
                    )
                )

                maximum_rows = min(
                    int(
                        formula_sheet.max_row
                        or 0
                    ),
                    spreadsheet_settings
                    .maximum_rows_per_sheet,
                    (
                        spreadsheet_settings
                        .maximum_total_rows
                        - total_rows_scanned
                    ),
                )

                maximum_columns = min(
                    int(
                        formula_sheet.max_column
                        or 0
                    ),
                    spreadsheet_settings
                    .maximum_columns,
                )

                if (
                    maximum_rows <= 0
                    or maximum_columns <= 0
                ):
                    continue

                original_max_row = int(
                    formula_sheet.max_row
                    or 0
                )

                original_max_column = int(
                    formula_sheet.max_column
                    or 0
                )

                sheet_truncated = (
                    original_max_row
                    > maximum_rows
                    or original_max_column
                    > maximum_columns
                )

                workbook_truncated = (
                    workbook_truncated
                    or sheet_truncated
                )

                headers: list[str] = []

                header_row_index: (
                    int
                    | None
                ) = None

                data_rows = 0
                sheet_rows_scanned = 0
                sheet_formula_count = 0
                duplicate_rows = 0

                missing_counts: dict[
                    int,
                    int,
                ] = defaultdict(int)

                numeric_stats: dict[
                    int,
                    dict[str, float],
                ] = {}

                row_hashes: set[
                    bytes
                ] = set()

                chunk_lines: list[str] = []
                chunk_start = 0
                chunk_end = 0

                def flush_chunk() -> None:
                    nonlocal chunk_lines
                    nonlocal chunk_start
                    nonlocal chunk_end

                    if not chunk_lines:
                        return

                    label = (
                        f"Sheet {safe_sheet_name} "
                        f"rows {chunk_start}-{chunk_end}"
                    )

                    text = "\n".join(
                        chunk_lines
                    )

                    sources.append(
                        SpreadsheetSource(
                            source_id=(
                                f"sheet-"
                                f"{hashlib.sha1(sheet_name.encode('utf-8')).hexdigest()[:10]}"
                                f"-rows-"
                                f"{chunk_start}-"
                                f"{chunk_end}"
                            ),
                            label=label,
                            kind="sheet_rows",
                            text=text,
                            character_count=len(
                                text
                            ),
                        )
                    )

                    chunk_lines = []
                    chunk_start = 0
                    chunk_end = 0

                formula_rows = (
                    formula_sheet.iter_rows(
                        min_row=1,
                        max_row=maximum_rows,
                        min_col=1,
                        max_col=maximum_columns,
                    )
                )

                value_rows = (
                    value_sheet.iter_rows(
                        min_row=1,
                        max_row=maximum_rows,
                        min_col=1,
                        max_col=maximum_columns,
                    )
                )

                for row_index, (
                    formula_row,
                    value_row,
                ) in enumerate(
                    zip(
                        formula_rows,
                        value_rows,
                    ),
                    start=1,
                ):
                    if (
                        monotonic()
                        - started_at
                        > spreadsheet_settings
                        .maximum_parse_seconds
                    ):
                        raise SpreadsheetValidationError(
                            "The workbook took too long to parse safely."
                        )

                    formula_values: list[
                        Any
                    ] = []

                    cached_values: list[
                        Any
                    ] = []

                    visible_values: list[str] = []

                    row_formula_count = 0

                    for (
                        formula_cell,
                        value_cell,
                    ) in zip(
                        formula_row,
                        value_row,
                    ):
                        formula_value = (
                            formula_cell.value
                        )

                        cached_value = (
                            value_cell.value
                        )

                        formula_values.append(
                            formula_value
                        )

                        cached_values.append(
                            cached_value
                        )

                        if (
                            formula_cell.data_type
                            == "f"
                        ):
                            row_formula_count += 1

                            formula_text = (
                                _cell_text(
                                    formula_value
                                )
                            )

                            cached_text = (
                                _cell_text(
                                    cached_value
                                )
                            )

                            if cached_text:
                                visible = (
                                    f"{formula_text} "
                                    f"=> {cached_text}"
                                )
                            else:
                                visible = (
                                    formula_text
                                )

                        else:
                            visible = _cell_text(
                                cached_value
                                if cached_value
                                is not None
                                else formula_value
                            )

                        visible_values.append(
                            visible
                        )

                    if not any(
                        visible_values
                    ):
                        continue

                    sheet_rows_scanned += 1
                    total_rows_scanned += 1

                    maximum_columns_seen = max(
                        maximum_columns_seen,
                        len(
                            visible_values
                        ),
                    )

                    sheet_formula_count += (
                        row_formula_count
                    )

                    if header_row_index is None:
                        header_row_index = (
                            row_index
                        )

                        headers = [
                            value
                            or (
                                "Column "
                                f"{get_column_letter(index)}"
                            )
                            for index, value in enumerate(
                                visible_values,
                                start=1,
                            )
                        ]

                        continue

                    data_rows += 1

                    while len(
                        headers
                    ) < len(
                        visible_values
                    ):
                        headers.append(
                            "Column "
                            f"{get_column_letter(len(headers) + 1)}"
                        )

                    signature = _row_signature(
                        visible_values[
                            :spreadsheet_settings
                            .maximum_context_columns
                        ]
                    )

                    if signature in row_hashes:
                        duplicate_rows += 1
                    else:
                        row_hashes.add(
                            signature
                        )

                    display_parts: list[
                        str
                    ] = []

                    for column_index, (
                        visible_value,
                        cached_value,
                    ) in enumerate(
                        zip(
                            visible_values,
                            cached_values,
                        )
                    ):
                        if column_index >= (
                            spreadsheet_settings
                            .maximum_context_columns
                        ):
                            break

                        header = headers[
                            column_index
                        ]

                        if not visible_value:
                            missing_counts[
                                column_index
                            ] += 1

                        if _is_numeric(
                            cached_value
                        ):
                            number = float(
                                cached_value
                            )

                            stats = (
                                numeric_stats
                                .setdefault(
                                    column_index,
                                    {
                                        "count": 0.0,
                                        "sum": 0.0,
                                        "min": number,
                                        "max": number,
                                    },
                                )
                            )

                            stats["count"] += 1
                            stats["sum"] += number
                            stats["min"] = min(
                                stats["min"],
                                number,
                            )
                            stats["max"] = max(
                                stats["max"],
                                number,
                            )

                        display_parts.append(
                            (
                                f"{header}="
                                f"{visible_value or '∅'}"
                            )
                        )

                    if not chunk_lines:
                        chunk_start = (
                            row_index
                        )

                    chunk_end = row_index

                    chunk_lines.append(
                        (
                            f"Row {row_index}: "
                            + "; ".join(
                                display_parts
                            )
                        )
                    )

                    if (
                        len(chunk_lines)
                        >= spreadsheet_settings
                        .row_chunk_size
                    ):
                        flush_chunk()

                        if (
                            len(sources)
                            >= spreadsheet_settings
                            .maximum_sources
                        ):
                            workbook_truncated = True
                            break

                    if (
                        total_rows_scanned
                        >= spreadsheet_settings
                        .maximum_total_rows
                    ):
                        workbook_truncated = True
                        break

                flush_chunk()

                total_formula_count += (
                    sheet_formula_count
                )

                if headers:
                    profile = _profile_text(
                        name=(
                            f"Sheet {safe_sheet_name}"
                        ),
                        headers=headers,
                        data_rows=data_rows,
                        columns_seen=min(
                            original_max_column,
                            spreadsheet_settings
                            .maximum_columns,
                        ),
                        formula_count=(
                            sheet_formula_count
                        ),
                        duplicate_rows=(
                            duplicate_rows
                        ),
                        missing_counts=dict(
                            missing_counts
                        ),
                        numeric_stats=(
                            numeric_stats
                        ),
                        truncated=(
                            sheet_truncated
                            or workbook_truncated
                        ),
                    )

                    sources.append(
                        SpreadsheetSource(
                            source_id=(
                                "sheet-"
                                f"{hashlib.sha1(sheet_name.encode('utf-8')).hexdigest()[:10]}"
                                "-profile"
                            ),
                            label=(
                                f"Sheet {safe_sheet_name} profile"
                            ),
                            kind="profile",
                            text=profile,
                            character_count=len(
                                profile
                            ),
                        )
                    )

                if (
                    total_rows_scanned
                    >= spreadsheet_settings
                    .maximum_total_rows
                    or len(sources)
                    >= spreadsheet_settings
                    .maximum_sources
                ):
                    workbook_truncated = True
                    break

            if not sources:
                raise SpreadsheetValidationError(
                    "The workbook does not contain readable spreadsheet data."
                )

            return ExtractedSpreadsheet(
                spreadsheet_type="xlsx",
                sheet_names=sheet_names,
                sources=tuple(
                    sources
                ),
                rows_scanned=(
                    total_rows_scanned
                ),
                maximum_columns_seen=(
                    maximum_columns_seen
                ),
                formula_count=(
                    total_formula_count
                ),
                truncated=(
                    workbook_truncated
                ),
            )

        finally:
            formula_workbook.close()
            value_workbook.close()

    @staticmethod
    def _extract_sync(
        *,
        file_bytes: bytes,
        extension: str,
    ) -> ExtractedSpreadsheet:
        if extension == ".csv":
            return (
                SpreadsheetService
                ._extract_csv_sync(
                    file_bytes
                )
            )

        if extension == ".xlsx":
            return (
                SpreadsheetService
                ._extract_xlsx_sync(
                    file_bytes
                )
            )

        raise SpreadsheetValidationError(
            "Unsupported spreadsheet type."
        )

    @staticmethod
    def _select_sources(
        spreadsheet: ExtractedSpreadsheet,
        prompt: str,
    ) -> tuple[
        SpreadsheetSource,
        ...
    ]:
        sources = spreadsheet.sources

        maximum = (
            spreadsheet_settings
            .maximum_selected_sources
        )

        if len(sources) <= maximum:
            return sources

        selected: dict[
            str,
            SpreadsheetSource,
        ] = {}

        def add_source(
            source: SpreadsheetSource,
        ) -> None:
            if len(selected) < maximum:
                selected.setdefault(
                    source.source_id,
                    source,
                )

        profiles = tuple(
            source
            for source in sources
            if source.kind == "profile"
        )

        row_sources = tuple(
            source
            for source in sources
            if source.kind != "profile"
        )

        for profile in profiles:
            add_source(
                profile
            )

        if _looks_like_broad_request(
            prompt
        ):
            remaining = (
                maximum
                - len(selected)
            )

            for source in _evenly_sample(
                row_sources,
                remaining,
            ):
                add_source(
                    source
                )

        else:
            terms = _query_terms(
                prompt
            )

            ranked = sorted(
                sources,
                key=lambda source: (
                    _source_score(
                        source,
                        terms,
                    ),
                    (
                        1
                        if source.kind
                        == "profile"
                        else 0
                    ),
                ),
                reverse=True,
            )

            for source in ranked:
                if (
                    _source_score(
                        source,
                        terms,
                    )
                    <= 0
                ):
                    continue

                add_source(
                    source
                )

            remaining = (
                maximum
                - len(selected)
            )

            if remaining > 0:
                for source in _evenly_sample(
                    row_sources,
                    remaining,
                ):
                    add_source(
                        source
                    )

        return tuple(
            source
            for source in sources
            if source.source_id
            in selected
        )

    @staticmethod
    def _build_context(
        sources: tuple[
            SpreadsheetSource,
            ...
        ],
        truncated: bool,
    ) -> str:
        parts = [
            (
                "WORKBOOK TRUNCATED DURING EXTRACTION: "
                f"{'yes' if truncated else 'no'}"
            )
        ]

        current_characters = len(
            parts[0]
        )

        for source in sources:
            section = (
                f"\n\n===== SOURCE: "
                f"{source.label} =====\n"
                f"{source.text}"
            )

            if (
                current_characters
                + len(section)
                > spreadsheet_settings
                .maximum_context_characters
            ):
                remaining = (
                    spreadsheet_settings
                    .maximum_context_characters
                    - current_characters
                )

                if remaining > 300:
                    parts.append(
                        section[:remaining]
                    )

                break

            parts.append(
                section
            )

            current_characters += len(
                section
            )

        return "".join(
            parts
        )

    @staticmethod
    def _canonicalize_citations(
        *,
        answer: str,
        sources: tuple[
            SpreadsheetSource,
            ...
        ],
    ) -> str:
        def replace_group(
            match: re.Match[str],
        ) -> str:
            content = (
                match.group(1)
                .strip()
                .lower()
            )

            labels: list[str] = []

            for source in sources:
                if (
                    source.label.lower()
                    in content
                    and source.label
                    not in labels
                ):
                    labels.append(
                        source.label
                    )

            if not labels:
                return match.group(0)

            return " ".join(
                f"[{label}]"
                for label in labels
            )

        return re.sub(
            r"\[([^\[\]\n]{1,240})\]",
            replace_group,
            answer,
        )

    @staticmethod
    def _extract_citations(
        *,
        answer: str,
        sources: tuple[
            SpreadsheetSource,
            ...
        ],
    ) -> tuple[
        SpreadsheetCitation,
        ...
    ]:
        lowered = answer.lower()

        citations: list[
            SpreadsheetCitation
        ] = []

        for source in sources:
            token = (
                f"[{source.label}]"
            )

            if token.lower() not in lowered:
                continue

            citations.append(
                SpreadsheetCitation(
                    source_id=(
                        source.source_id
                    ),
                    label=source.label,
                    kind=source.kind,
                )
            )

        return tuple(
            citations
        )

    @staticmethod
    def _response_text(
        response: Any,
    ) -> str:
        try:
            return (
                getattr(
                    response,
                    "text",
                    "",
                )
                or ""
            ).strip()

        except Exception:
            return ""

    @staticmethod
    def _usage(
        response: Any,
    ) -> TokenUsage:
        metadata = getattr(
            response,
            "usage_metadata",
            None,
        )

        if metadata is None:
            return TokenUsage()

        prompt_tokens = int(
            getattr(
                metadata,
                "prompt_token_count",
                0,
            )
            or 0
        )

        completion_tokens = int(
            getattr(
                metadata,
                "candidates_token_count",
                0,
            )
            or getattr(
                metadata,
                "output_token_count",
                0,
            )
            or 0
        )

        total_tokens = int(
            getattr(
                metadata,
                "total_token_count",
                0,
            )
            or (
                prompt_tokens
                + completion_tokens
            )
        )

        return TokenUsage(
            prompt_tokens=prompt_tokens,
            completion_tokens=(
                completion_tokens
            ),
            total_tokens=total_tokens,
        )

    @staticmethod
    def _status_code(
        error: Exception,
    ) -> int | None:
        try:
            return int(
                getattr(
                    error,
                    "code",
                    None,
                )
            )

        except (
            TypeError,
            ValueError,
        ):
            return None

    async def analyze(
        self,
        *,
        file_bytes: bytes,
        extension: str,
        prompt: str,
    ) -> SpreadsheetAnalysis:
        spreadsheet = await asyncio.to_thread(
            self._extract_sync,
            file_bytes=file_bytes,
            extension=extension,
        )

        selected_sources = (
            self._select_sources(
                spreadsheet,
                prompt,
            )
        )

        context = self._build_context(
            selected_sources,
            spreadsheet.truncated,
        )

        labels = tuple(
            source.label
            for source in selected_sources
        )

        request_content = f"""
USER REQUEST

{prompt}

SPREADSHEET TYPE

{spreadsheet.spreadsheet_type}

SUPPLIED SOURCE LABELS

{", ".join(labels)}

SPREADSHEET SOURCES

{context}

Answer using only the supplied spreadsheet sources.
Use exact inline citations such as
[Sheet Revenue profile],
[Sheet Revenue rows 2-20], or
[CSV rows 2-20].
""".strip()

        last_error: (
            Exception
            | None
        ) = None

        for model in (
            self._model_candidates()
        ):
            try:
                response = (
                    await self
                    ._get_client()
                    .aio
                    .models
                    .generate_content(
                        model=model,
                        contents=request_content,
                        config=types.GenerateContentConfig(
                            system_instruction=(
                                SPREADSHEET_SYSTEM_PROMPT
                            ),
                            temperature=0.1,
                            max_output_tokens=3_000,
                        ),
                    )
                )

                answer = (
                    self._response_text(
                        response
                    )
                )

                if not answer:
                    raise SpreadsheetResponseError(
                        "The spreadsheet model returned an empty answer."
                    )

                answer = (
                    self._canonicalize_citations(
                        answer=answer,
                        sources=selected_sources,
                    )
                )

                citations = (
                    self._extract_citations(
                        answer=answer,
                        sources=selected_sources,
                    )
                )

                if not citations:
                    reviewed = " ".join(
                        f"[{source.label}]"
                        for source
                        in selected_sources[:6]
                    )

                    answer = (
                        answer.rstrip()
                        + "\n\n"
                        + "**Sources reviewed:** "
                        + reviewed
                    )

                    citations = tuple(
                        SpreadsheetCitation(
                            source_id=(
                                source.source_id
                            ),
                            label=source.label,
                            kind=source.kind,
                        )
                        for source
                        in selected_sources[:6]
                    )

                return SpreadsheetAnalysis(
                    answer=answer,
                    model=model,
                    request_id=getattr(
                        response,
                        "response_id",
                        None,
                    ),
                    usage=self._usage(
                        response
                    ),
                    spreadsheet_type=(
                        spreadsheet
                        .spreadsheet_type
                    ),
                    sheet_names=(
                        spreadsheet
                        .sheet_names
                    ),
                    rows_scanned=(
                        spreadsheet
                        .rows_scanned
                    ),
                    maximum_columns_seen=(
                        spreadsheet
                        .maximum_columns_seen
                    ),
                    formula_count=(
                        spreadsheet
                        .formula_count
                    ),
                    truncated=(
                        spreadsheet
                        .truncated
                    ),
                    selected_sources=(
                        labels
                    ),
                    citations=citations,
                )

            except Exception as error:
                last_error = error

                logger.warning(
                    "Spreadsheet model failed: "
                    "model=%s type=%s status=%s",
                    model,
                    type(error).__name__,
                    self._status_code(
                        error
                    ),
                )

                if isinstance(
                    error,
                    errors.APIError,
                ):
                    status_code = (
                        self._status_code(
                            error
                        )
                    )

                    if status_code in {
                        401,
                        403,
                    }:
                        raise SpreadsheetConfigurationError(
                            "Gemini spreadsheet credentials "
                            "are invalid or unauthorized."
                        ) from error

                    if status_code not in {
                        404,
                        408,
                        429,
                        500,
                        502,
                        503,
                        504,
                    }:
                        raise SpreadsheetResponseError(
                            "Gemini rejected the spreadsheet request."
                        ) from error

                elif not isinstance(
                    error,
                    SpreadsheetResponseError,
                ):
                    raise SpreadsheetResponseError(
                        "The spreadsheet request could not be completed."
                    ) from error

        raise SpreadsheetResponseError(
            "No configured model produced a usable spreadsheet answer."
        ) from last_error
